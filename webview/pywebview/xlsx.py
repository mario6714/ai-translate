from typing import Dict, List, Tuple, Optional, Union, Any
import ast
import openpyxl, os, csv, re
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment



file_name: Optional[str] = None
workbook: Optional[Workbook] = None

def get_save_dir():
    env = os.getenv('appdata')
    if env:
        path = os.path.join(env, 'ai-translate')
        if not os.path.exists(path): os.mkdir(path)
        return path

def filePath(): 
    save_dir = get_save_dir()
    if save_dir and file_name:
        xlsx_path = os.path.join(save_dir, file_name+'.xlsx')
        csv_path = os.path.join(save_dir, file_name+'.csv')
        if os.path.isfile(csv_path) and not os.path.exists(xlsx_path): 
            return csvToXLSX(csv_path, xlsx_path)
        return xlsx_path

def csvToXLSX(csv_path: str, xlsx_path: str) -> str:
    if isinstance(csv_path, str) and isinstance(xlsx_path, str):
        wb = Workbook()
        sheet = wb.active
        if not sheet: raise Exception("CSV convertion error: newly created workbook has no sheets!")
        sheet.title = 'Translation'

        with open(csv_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader: 
                if '' not in row[0]: sheet.append(row)

        wb.save(xlsx_path)
        return xlsx_path


class CustomComment(Comment):
    def __init__(self, speaker_name: str, src_model: str):
        data = {}
        if self._validateInput(speaker_name): data['speaker_name'] = speaker_name
        if self._validateInput(src_model): data['src_model'] = src_model
        super().__init__(str(data), author= "")

    def _validateInput(self, text: str) -> bool: 
        if isinstance(text, str) and len(text): return True
        return False

    @staticmethod
    def get_speaker_name(text: str) -> str: 
        try:
            data = ast.literal_eval(text)
            return data['speaker_name']
        except: return ""


class TextDTO:
    history: Optional[List[str]]
    translatedText: Union[None, str, List[Any]]
    originalText: str
    def __init__(self, DTO: Dict[str, str]):
        self.window_title = DTO["window_title"]
        self.originalText = DTO["originalText"]
        self.src_model = DTO['src_model'] if 'src_model' in DTO else None
        self.translatedText = DTO["translatedText"] if "translatedText" in DTO else None
        self._speakerName = DTO['speakerName'] if 'speakerName' in DTO else None

    @property
    def speakerName(self):
        return self._speakerName if self._speakerName is not None else ""


def loadFile(name: str): 
    global workbook, file_name
    file_name = name
    file_path = filePath()
    if file_path: workbook = openpyxl.load_workbook(file_path)
    else: 
        workbook = Workbook()
        sheet = workbook.active
        if not sheet: raise Exception("Newly create workbook has no sheets!")
        sheet.title = 'Translation'

def worksheet() -> Worksheet: # type: ignore
    if workbook is not None: return workbook['Translation']


def queryEntry(textDTO: TextDTO):
    if isinstance(textDTO, TextDTO) and textDTO.window_title is not None:
        if textDTO.window_title != file_name or worksheet() is None: loadFile(textDTO.window_title)
        if worksheet() is not None:
            column: Tuple[Cell] = worksheet()['A']
            for cell in column:
                value = cell.value
                if isinstance(value, str):
                    value = re.sub(r"\[★\]", "", value)
                    if textDTO.originalText == value: return cell.row


def get_history(lastRowNumber: int) -> list[Any]:
    if isinstance(lastRowNumber, int):
        history = []
        rowNumber = lastRowNumber
        while len(history) < 10:
            rowNumber = rowNumber-1
            if rowNumber > 0:
                cell = worksheet().cell(row= rowNumber, column=2)
                if cell.value is not None and cell.value != "": 
                    speaker_name = CustomComment.get_speaker_name(cell.comment.text) if cell.comment is not None else ""
                    formated_name = f"[{speaker_name}]: " if len(speaker_name) else ""
                    if isinstance(cell.value, str): history.append(formated_name + cell.value)
            else: break

        history.reverse()
        return history



class XLSX:
    def QueryTranslation(self, text_dto_dict: Dict[str, Any]) -> dict[str, Any]: 
        textDTO: TextDTO = TextDTO(text_dto_dict)
        entry = queryEntry(textDTO)
        if entry is not None: 
            textDTO.history = get_history(entry)
            textDTO.translatedText = [cell.value for cell in worksheet()[entry][1:]]
            textDTO.translatedText.reverse()
        elif worksheet() is not None: textDTO.history = get_history(worksheet().max_row+1)

        return textDTO.__dict__


    def SaveText(self, text_dto_dict: Dict[str, Any]):
        textDTO: TextDTO = TextDTO(text_dto_dict)
        entry = queryEntry(textDTO)
        if entry is not None and worksheet() is not None: 
            row = worksheet()[entry]
            cell = worksheet().cell(row= entry, column= len(row))
            cell.value = textDTO.translatedText # type: ignore
            #cell.comment = CustomComment(textDTO.speakerName, textDTO.src_model)

        elif worksheet() is not None: 
            lastRow = worksheet().max_row+1 # same variable for both calls to avoid the "stairs" bug
            cellA = worksheet().cell(row= lastRow, column=1)
            cellB = worksheet().cell(row= lastRow, column=2)
            cellA.value = textDTO.originalText # type: ignore
            cellB.value = textDTO.translatedText # type: ignore
            #cellB.comment = CustomComment(textDTO.speakerName, textDTO.src_model)

        else: 
            return { "error": "failed to save text" }

        file_path = filePath()
        if not workbook or not file_path: raise Exception('Failed to save workbook: workbook or file not found!')
        workbook.save(file_path)


